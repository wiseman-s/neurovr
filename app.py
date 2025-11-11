# app.py — NeuroVR Lab v3.5 (full)
# AI + VR + Drug Discovery + Research Dashboard + Patient Recorder (session-only)
# System by Simon | allinmer57@gmail.com

import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import joblib, shap, base64, io, os, matplotlib.pyplot as plt
from datetime import datetime
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from streamlit.components.v1 import html

# --------------------------- CONFIG ---------------------------
st.set_page_config(
    page_title=" NeuroVR Lab — VR Stroke Research & Drug Discovery",
    page_icon="🧠",
    layout="wide"
)

# --------------------------- SESSION STATE SETUP ---------------------------
if "patient_records" not in st.session_state:
    st.session_state.patient_records = pd.DataFrame(
        columns=["id", "age", "bp", "cholesterol", "glucose", "bmi", "stroke", "timestamp"]
    )

if "compounds" not in st.session_state:
    st.session_state.compounds = []  # store evaluated compounds in-session

# fixed home notes (your preferred text)
if "notes" not in st.session_state:
    st.session_state.notes = (
        "🧠 **NeuroVR Lab: AI-Powered Stroke & Drug Discovery Platform**\n\n"
        "Condition: Stroke (Cerebrovascular Accident) is a major neurological disorder caused by interruption "
        "of blood flow to the brain, leading to tissue damage.\n\n"
        "Goal: NeuroVR Lab combines AI prediction, patient data simulation, and virtual reality visualization "
        "to enhance understanding of stroke patterns and aid discovery of potential therapeutic compounds.\n\n"
        "Impact: This system offers an experimental framework for early risk detection, drug interaction testing, "
        "and educational visualization."
    )

# --------------------------- HELPERS ---------------------------
def train_model(dataframe):
    # Expect dataframe with columns: age,bp,cholesterol,glucose,bmi,stroke
    X = dataframe[["age", "bp", "cholesterol", "glucose", "bmi"]]
    y = dataframe["stroke"]
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)
    model_local = RandomForestClassifier(n_estimators=100, random_state=42)
    model_local.fit(X_train, y_train)
    acc_local = accuracy_score(y_test, model_local.predict(X_test))
    return model_local, acc_local

def safe_shap_plot(model, X):
    """
    Compute SHAP values and return a matplotlib Figure and the shap array
    normalized into a 2D array (samples x features). This avoids shape issues
    that lead to 'Per-column arrays must each be 1-dimensional' errors.
    """
    explainer = shap.TreeExplainer(model)
    shap_values = explainer.shap_values(X)

    # For binary classification, shap_values can be a list [neg, pos]
    shap_to_plot = shap_values[1] if isinstance(shap_values, list) else shap_values
    shap_to_plot = np.array(shap_to_plot)

    # If shap_to_plot has extra dimensions collapse them
    if shap_to_plot.ndim > 2:
        shap_to_plot = shap_to_plot.reshape(shap_to_plot.shape[0], -1)

    # Make sure shap columns align with X columns
    if shap_to_plot.shape[1] != X.shape[1]:
        # Trim or pad with zeros to match feature count
        if shap_to_plot.shape[1] > X.shape[1]:
            shap_to_plot = shap_to_plot[:, :X.shape[1]]
        else:
            pad_width = X.shape[1] - shap_to_plot.shape[1]
            shap_to_plot = np.pad(shap_to_plot, ((0,0),(0,pad_width)), mode="constant", constant_values=0.0)

    # Draw summary_plot onto a matplotlib figure
    plt.figure(figsize=(8, 5))
    shap.summary_plot(shap_to_plot, X, show=False)
    fig = plt.gcf()
    plt.close(fig)
    return fig, shap_to_plot

def add_patient_record(age, bp, chol, glucose, bmi, stroke):
    df = st.session_state.patient_records
    new_id = int(df["id"].max()) + 1 if not df.empty else 1
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    new_row = {"id": new_id, "age": age, "bp": bp, "cholesterol": chol,
               "glucose": glucose, "bmi": bmi, "stroke": stroke, "timestamp": ts}
    st.session_state.patient_records = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)

def delete_patient_record(record_id):
    st.session_state.patient_records = st.session_state.patient_records[st.session_state.patient_records["id"] != record_id].reset_index(drop=True)

# --------------------------- LOAD SYNTHETIC DATA & INITIAL MODEL ---------------------------
@st.cache_data
def load_synthetic():
    np.random.seed(42)
    return pd.DataFrame({
        "age": np.random.randint(30, 85, 200),
        "bp": np.random.randint(110, 190, 200),
        "cholesterol": np.random.randint(100, 300, 200),
        "glucose": np.random.randint(60, 200, 200),
        "bmi": np.random.uniform(18, 35, 200),
        "stroke": np.random.choice([0, 1], 200, p=[0.75, 0.25])
    })

data = load_synthetic()
model, acc = train_model(data)

# --------------------------- SIDEBAR NAVIGATION ---------------------------
st.sidebar.header("Navigation")
page = st.sidebar.radio("Go to:", [
    "Home 🏠",
    "Research Dashboard 📊",
    "AI Stroke Predictor 🤖",
    "Patient Data Recorder 🧾",
    "Drug Discovery Lab 💊",
    "VR Headset Mode 🕶️",
    "Upload/Train Dataset 📂",
    "Research Report 🧾"
])

# --------------------------- HOME ---------------------------
if page == "Home 🏠":
    st.markdown("<h2 style='color:#1f77b4;'>🧬 NeuroVR Lab — VR Stroke Research & Drug Discovery</h2>", unsafe_allow_html=True)
    st.info(f"Current AI Model Accuracy: {acc*100:.2f}%")

    st.write("### Overview & Research Focus")
    st.markdown(st.session_state.notes)

    st.write("---")
    st.write("### Sample Dataset Preview")
    st.dataframe(data.head())

    st.write("### 3D Stroke Risk Visualization")
    fig3d = px.scatter_3d(data, x="age", y="bp", z="cholesterol", color="stroke",
                         title="3D Stroke Risk Distribution", opacity=0.7)
    st.plotly_chart(fig3d, use_container_width=True)

# --------------------------- RESEARCH DASHBOARD ---------------------------
elif page == "Research Dashboard 📊":
    st.subheader("Research Dashboard: Data Insights")
    st.markdown("Interactive charts update with uploaded dataset or session patient records.")

    # choose data source: base synthetic, uploaded/train, or session records
    source = st.selectbox("Data source for dashboard:", ["Synthetic dataset", "Patient session records"])
    if source == "Patient session records" and not st.session_state.patient_records.empty:
        ds = st.session_state.patient_records.copy()
        # convert types if needed
        ds[["age","bp","cholesterol","glucose","bmi","stroke"]] = ds[["age","bp","cholesterol","glucose","bmi","stroke"]].apply(pd.to_numeric, errors="coerce")
    else:
        ds = data.copy()

    col1, col2 = st.columns(2)
    with col1:
        # handle when stroke column is missing or non-numeric
        if "stroke" in ds.columns and pd.api.types.is_numeric_dtype(ds["stroke"]):
            color_series = ds["stroke"].map({0: "No", 1: "Yes"})
        else:
            color_series = None
        fig_age = px.histogram(ds, x="age", color=color_series, barmode="overlay", title="Age Distribution vs Stroke")
        st.plotly_chart(fig_age, use_container_width=True)
    with col2:
        fig_bp = px.scatter(ds, x="bp", y="cholesterol", color=color_series, title="BP vs Cholesterol")
        st.plotly_chart(fig_bp, use_container_width=True)

    st.markdown("### Correlation Map")
    numeric = ds.select_dtypes(include=[np.number])
    if not numeric.empty:
        fig_corr = px.imshow(numeric.corr(), text_auto=True, aspect="auto", title="Feature Correlation Map")
        st.plotly_chart(fig_corr, use_container_width=True)

    st.markdown("### Quick Insights")
    if "stroke" in ds.columns and ds["stroke"].dtype != object and ds["stroke"].notna().any():
        avg_age_stroke = ds[ds["stroke"] == 1]["age"].mean()
        avg_bp_stroke = ds[ds["stroke"] == 1]["bp"].mean()
        st.write(f"- Average age among stroke cases: **{avg_age_stroke:.1f}**")
        st.write(f"- Average BP among stroke cases: **{avg_bp_stroke:.1f}**")
    else:
        st.info("No stroke-labeled records in current view.")

# --------------------------- AI STROKE PREDICTOR ---------------------------
elif page == "AI Stroke Predictor 🤖":
    st.subheader("AI Stroke Risk Prediction (Manual entry or session record)")

    use_record = st.checkbox("Use a patient from session records", value=False)
    if use_record and not st.session_state.patient_records.empty:
        sel = st.selectbox("Select patient ID", options=list(st.session_state.patient_records["id"]))
        patient_row = st.session_state.patient_records[st.session_state.patient_records["id"] == sel].iloc[0]
        age = int(patient_row["age"]); bp = float(patient_row["bp"]); chol = float(patient_row["cholesterol"])
        glucose = float(patient_row["glucose"]); bmi = float(patient_row["bmi"])
        st.write(f"Using patient ID {sel} — Age {age}, BP {bp}, Chol {chol}")
    else:
        age = st.number_input("Age", 18, 100, 45)
        bp = st.number_input("Blood Pressure", 60, 240, 120)
        chol = st.number_input("Cholesterol", 50, 400, 180)
        glucose = st.number_input("Glucose", 40, 400, 100)
        bmi = st.number_input("BMI", 10.0, 60.0, 25.0)

    if st.button("Predict Stroke Risk"):
        X_input = pd.DataFrame([[age, bp, chol, glucose, bmi]], columns=["age","bp","cholesterol","glucose","bmi"])
        prob = model.predict_proba(X_input)[0][1]
        pred = model.predict(X_input)[0]
        st.success(f"Predicted stroke probability: **{prob*100:.2f}%**")
        if pred == 1:
            st.warning("High risk — recommend clinical follow-up.")
        else:
            st.info("Low risk — maintain healthy lifestyle.")

        # SHAP explainability (safe)
        X_full = data[["age","bp","cholesterol","glucose","bmi"]]
        try:
            fig_shap, shap_to_plot = safe_shap_plot(model, X_full)
            st.pyplot(fig_shap)
            # feature importance
            shap_values_mean = np.abs(shap_to_plot).mean(axis=0)
            shap_values_mean = np.array(shap_values_mean).ravel()
            feature_count = X_full.shape[1]
            if len(shap_values_mean) > feature_count:
                shap_values_mean = shap_values_mean[:feature_count]
            elif len(shap_values_mean) < feature_count:
                shap_values_mean = np.pad(shap_values_mean, (0, feature_count - len(shap_values_mean)), 'constant')
            importance_df = pd.DataFrame({"Feature": X_full.columns, "Importance": shap_values_mean}).sort_values("Importance", ascending=False)
            st.plotly_chart(px.bar(importance_df, x="Feature", y="Importance", title="Feature importance (SHAP)"), use_container_width=True)
        except Exception as e:
            st.error(f"SHAP visualization failed: {e}")

# --------------------------- PATIENT DATA RECORDER ---------------------------
elif page == "Patient Data Recorder 🧾":
    st.subheader("Patient Data Recorder (session-only)")

    with st.form("add_patient", clear_on_submit=True):
        col1, col2, col3 = st.columns(3)
        with col1:
            age = st.number_input("Age", 0, 120, 45, key="rec_age")
            bp = st.number_input("Blood Pressure", 40, 300, 120, key="rec_bp")
            chol = st.number_input("Cholesterol", 0, 600, 180, key="rec_chol")
        with col2:
            glucose = st.number_input("Glucose", 0, 600, 100, key="rec_gluc")
            bmi = st.number_input("BMI", 5.0, 80.0, 25.0, key="rec_bmi")
            stroke = st.selectbox("Stroke (0=no,1=yes)", [0,1], index=0, key="rec_stroke")
        with col3:
            submitted = st.form_submit_button("Add Patient Record")
            if submitted:
                add_patient_record(age, bp, chol, glucose, bmi, stroke)
                st.success("Patient record added (session-only).")

    st.write("### Session Patient Records")
    df = st.session_state.patient_records.copy()
    if df.empty:
        st.info("No session records yet — add patients above.")
    else:
        search = st.text_input("Search by ID (leave blank to show all)")
        if search:
            try:
                sid = int(search)
                df = df[df["id"] == sid]
            except:
                st.warning("Enter a numeric ID to search.")
        st.dataframe(df)

        # delete UI
        st.markdown("#### Delete a record")
        if not df.empty:
            delete_id = st.number_input("Enter ID to delete", min_value=int(df["id"].min()), max_value=int(df["id"].max()), step=1)
            if st.button("Delete Record"):
                delete_patient_record(delete_id)
                st.success(f"Record {delete_id} deleted from session.")

# --------------------------- DRUG DISCOVERY LAB ---------------------------
elif page == "Drug Discovery Lab 💊":
    st.subheader("Drug Discovery Lab — ML Surrogate Compound Efficacy & Risk Visualizer")

    st.markdown("""
    This lab evaluates compounds using a small ML surrogate trained on a synthetic,
    biologically-informed dataset (fast, in-memory). Use Manual Comparison for two compounds
    or Batch Upload (CSV with columns: Compound, Binding, Solubility, Toxicity).
    """)

    import io
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    from sklearn.ensemble import RandomForestRegressor
    from sklearn.model_selection import train_test_split

    # ---------------- ML SURROGATE (train once, cached) ----------------
    @st.cache_resource
    def build_surrogate(seed=42):
        rng = np.random.default_rng(seed)
        n = 2000
        # binding range: -15 (strong) to -1 (weak)
        binding = rng.uniform(-15, -1, size=n)
        sol = rng.uniform(0.0, 1.0, size=n)
        tox = rng.uniform(0.0, 1.0, size=n)

        # create a plausible target using domain-informed formula + noise
        binding_strength = np.clip((-binding - 5) / 7, 0.0, 1.0)   # maps -5->0, -12->1
        raw_score = 0.55 * binding_strength + 0.3 * sol + 0.15 * (1 - tox)
        noise = rng.normal(0, 0.03, size=n)   # small noise
        efficacy = np.clip((raw_score + noise) * 100, 0, 100)

        X = np.vstack([binding, sol, tox]).T
        y = efficacy

        X_train, X_val, y_train, y_val = train_test_split(X, y, test_size=0.2, random_state=seed)
        model = RandomForestRegressor(n_estimators=150, random_state=seed)
        model.fit(X_train, y_train)
        val_score = model.score(X_val, y_val)  # R^2 on validation
        return model, val_score

    surrogate_model, surrogate_r2 = build_surrogate()

    # ---------------- helper: evaluate via ML surrogate ----------------
    def evaluate_compound_ml(binding, solubility, toxicity):
        """
        Use ML surrogate to predict efficacy (0-100), stroke reduction (0-50),
        a confidence score (5-100) and reasons (list).
        """
        reasons = []
        # basic input validation / clipping
        binding = float(binding)
        sol = float(solubility)
        tox = float(toxicity)

        # domain heuristics for textual reasons
        if binding > -5:
            reasons.append("Weak binding (binding energy > -5 kcal/mol).")
        if sol < 0.25:
            reasons.append("Low solubility — bioavailability may be poor.")
        if tox > 0.6:
            reasons.append("High predicted toxicity — safety concerns.")

        # predict with surrogate model
        X_in = np.array([[binding, sol, tox]])
        pred_eff = surrogate_model.predict(X_in)[0]
        pred_eff = float(np.clip(pred_eff, 0.0, 100.0))
        pred_reduction = round((pred_eff / 100.0) * 50.0, 2)

        # confidence heuristic:
        # combine model validation R^2 and how in-distribution the inputs are
        # compute simple distance to mid-range of training distribution (heuristic)
        # binding_strength used to indicate how "typical" the binding is
        binding_strength = np.clip((-binding - 5) / 7, 0.0, 1.0)
        in_range_score = 1.0
        # penalize extremes
        if sol < 0.05 or sol > 0.98:
            in_range_score -= 0.2
        if tox < 0.0 or tox > 1.0:
            in_range_score -= 0.2
        # combine
        conf = surrogate_r2 * 60 + (binding_strength * 20) + (in_range_score * 20)
        conf = int(np.clip(conf, 5, 100))

        return round(pred_eff, 2), pred_reduction, conf, reasons

    # ---------------- Excel export helper ----------------
    def export_excel(df):
        wb = Workbook()
        ws = wb.active
        ws.title = "Drug Discovery Results"
        ws.append(list(df.columns))
        for _, row in df.iterrows():
            ws.append(row.tolist())
            # color by efficacy/toxicity
            eff = float(row.get("Efficacy", 0.0))
            tox = float(row.get("Toxicity", 0.0))
            if eff > 75 and tox < 0.3:
                color = "90EE90"  # light green
            elif eff < 40:
                color = "FFA07A"  # orange
            else:
                color = "FFFACD"  # light yellow
            for cell in ws[ws.max_row]:
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # ---------------- UI: mode selector ----------------
    mode = st.radio("Select Mode", ["Manual Comparison", "Batch Upload (CSV)", "Manual Multi-Entry"])

    # reference (real) compounds with approximate properties
    compound_db = {
        "Aspirin": {"binding": -8.0, "solubility": 0.7, "toxicity": 0.2},
        "Clopidogrel": {"binding": -9.0, "solubility": 0.5, "toxicity": 0.25},
        "Atorvastatin": {"binding": -10.0, "solubility": 0.4, "toxicity": 0.25},
        "Citicoline": {"binding": -7.2, "solubility": 0.8, "toxicity": 0.1},
        "Edaravone": {"binding": -11.0, "solubility": 0.6, "toxicity": 0.15},
        "tPA": {"binding": -13.5, "solubility": 0.45, "toxicity": 0.35},
        "Water (control)": {"binding": -1.0, "solubility": 1.0, "toxicity": 0.0},
        "Custom": {"binding": -8.0, "solubility": 0.5, "toxicity": 0.2},
    }

    compound_records = []  # will collect evaluated results

    # ---------------- Manual Comparison (two compounds) ----------------
    if mode == "Manual Comparison":
        colA, colB = st.columns(2)
        with colA:
            st.markdown("##### Compound A")
            name_a = st.selectbox("Select Compound A", list(compound_db.keys()), index=0, key="cmp_a")
            if name_a != "Custom":
                d = compound_db[name_a]
                bind_a, sol_a, tox_a = d["binding"], d["solubility"], d["toxicity"]
                st.info(f"Auto-filled: Binding={bind_a}, Solubility={sol_a}, Toxicity={tox_a}")
            else:
                bind_a = st.number_input("Binding Energy (A) [kcal/mol]", -20.0, 0.0, -8.0, step=0.1)
                sol_a = st.slider("Solubility (A)", 0.0, 1.0, 0.5)
                tox_a = st.slider("Toxicity (A)", 0.0, 1.0, 0.2)
        with colB:
            st.markdown("##### Compound B")
            name_b = st.selectbox("Select Compound B", list(compound_db.keys()), index=1, key="cmp_b")
            if name_b != "Custom":
                d = compound_db[name_b]
                bind_b, sol_b, tox_b = d["binding"], d["solubility"], d["toxicity"]
                st.info(f"Auto-filled: Binding={bind_b}, Solubility={sol_b}, Toxicity={tox_b}")
            else:
                bind_b = st.number_input("Binding Energy (B) [kcal/mol]", -20.0, 0.0, -10.0, step=0.1)
                sol_b = st.slider("Solubility (B)", 0.0, 1.0, 0.6)
                tox_b = st.slider("Toxicity (B)", 0.0, 1.0, 0.1)

        if st.button("Run Compound Evaluation"):
            a_eff, a_red, a_conf, a_reasons = evaluate_compound_ml(bind_a, sol_a, tox_a)
            b_eff, b_red, b_conf, b_reasons = evaluate_compound_ml(bind_b, sol_b, tox_b)

            compound_records = [
                {"Compound": name_a, "Binding": bind_a, "Solubility": sol_a, "Toxicity": tox_a,
                 "Efficacy": a_eff, "StrokeReduction": a_red, "Confidence": a_conf},
                {"Compound": name_b, "Binding": bind_b, "Solubility": sol_b, "Toxicity": tox_b,
                 "Efficacy": b_eff, "StrokeReduction": b_red, "Confidence": b_conf},
            ]
            df_cmp = pd.DataFrame(compound_records)
            st.dataframe(df_cmp, use_container_width=True)

            # show metrics side-by-side
            c1, c2 = st.columns(2)
            with c1:
                st.markdown(f"### {name_a}")
                st.metric("Efficacy", f"{a_eff}/100")
                st.metric("Predicted Stroke Reduction", f"{a_red}%")
                st.metric("Confidence", f"{a_conf}/100")
                if a_reasons:
                    st.warning(" • " + "\n • ".join(a_reasons))
            with c2:
                st.markdown(f"### {name_b}")
                st.metric("Efficacy", f"{b_eff}/100")
                st.metric("Predicted Stroke Reduction", f"{b_red}%")
                st.metric("Confidence", f"{b_conf}/100")
                if b_reasons:
                    st.warning(" • " + "\n • ".join(b_reasons))

            # 3D visualization
            fig3d = px.scatter_3d(
                df_cmp, x="Binding", y="Solubility", z="Toxicity",
                color="StrokeReduction", symbol="Compound", size="Efficacy",
                color_continuous_scale="RdYlGn", title="3D Visualization: Efficacy vs Properties"
            )
            st.plotly_chart(fig3d, use_container_width=True)

            # 2D comparison
            st.plotly_chart(px.bar(df_cmp, x="Compound", y=["Efficacy", "StrokeReduction", "Confidence"], barmode="group"))

            # exports
            csv = df_cmp.to_csv(index=False).encode("utf-8")
            excel = export_excel(df_cmp)
            st.download_button("📥 Download Results (CSV)", csv, "compound_results.csv", "text/csv")
            st.download_button("📘 Download Results (Excel)", excel, "compound_results.xlsx")

    # ---------------- Batch CSV upload ----------------
    elif mode == "Batch Upload (CSV)":
        st.markdown("### Upload Compound Dataset (CSV)")
        uploaded = st.file_uploader("Upload file", type=["csv"])
        if uploaded is not None:
            df = pd.read_csv(uploaded)
            expected_cols = {"Compound", "Binding", "Solubility", "Toxicity"}
            if not expected_cols.issubset(df.columns):
                st.error(f"CSV must contain: {', '.join(expected_cols)}")
            else:
                def eval_row(row):
                    eff, red, conf, reasons = evaluate_compound_ml(row["Binding"], row["Solubility"], row["Toxicity"])
                    return pd.Series([eff, red, conf])
                df[["Efficacy", "StrokeReduction", "Confidence"]] = df.apply(eval_row, axis=1)
                st.success(f"✅ {len(df)} compounds evaluated successfully.")
                st.dataframe(df.head(), use_container_width=True)
                fig3d = px.scatter_3d(
                    df, x="Binding", y="Solubility", z="Toxicity",
                    color="StrokeReduction", size="Efficacy", color_continuous_scale="RdYlGn",
                    title="3D Screening: Compound Properties & Stroke Reduction Potential"
                )
                st.plotly_chart(fig3d, use_container_width=True)
                st.plotly_chart(px.bar(df.sort_values("Efficacy", ascending=False).head(10), x="Compound", y="Efficacy", color="StrokeReduction"))
                st.write("### Summary Statistics")
                st.write(df[["Efficacy", "StrokeReduction", "Confidence"]].describe().T)
                best = df.loc[df["Efficacy"].idxmax()]
                st.success(f"🏆 Best Compound: **{best['Compound']}** (Efficacy {best['Efficacy']:.1f}, Stroke Reduction {best['StrokeReduction']:.1f}%)")
                csv = df.to_csv(index=False).encode("utf-8")
                excel = export_excel(df)
                st.download_button("📊 Download Full Results (CSV)", csv, "batch_results.csv", "text/csv")
                st.download_button("📘 Download Full Results (Excel)", excel, "batch_results.xlsx")

    # ---------------- Manual multi-entry (quick table entry) ----------------
    else:
        st.markdown("### Manual Multi-Entry — add up to 10 custom compounds")
        n = st.number_input("How many compounds?", 1, 10, 3)
        rows = []
        for i in range(int(n)):
            st.markdown(f"**Compound {i+1}**")
            cname = st.text_input(f"Name {i+1}", f"Compound_{i+1}", key=f"name_{i}")
            cbinding = st.number_input(f"Binding (kcal/mol) {i+1}", -20.0, 0.0, -8.0, step=0.1, key=f"bind_{i}")
            csol = st.slider(f"Solubility (0-1) {i+1}", 0.0, 1.0, 0.5, key=f"sol_{i}")
            ctox = st.slider(f"Toxicity (0-1) {i+1}", 0.0, 1.0, 0.2, key=f"tox_{i}")
            eff, red, conf, reasons = evaluate_compound_ml(cbinding, csol, ctox)
            rows.append({"Compound": cname, "Binding": cbinding, "Solubility": csol, "Toxicity": ctox,
                         "Efficacy": eff, "StrokeReduction": red, "Confidence": conf})
        if st.button("Analyze Manual Entries"):
            df_manual = pd.DataFrame(rows)
            st.dataframe(df_manual, use_container_width=True)
            st.plotly_chart(px.scatter_3d(df_manual, x="Binding", y="Solubility", z="Toxicity",
                                          color="StrokeReduction", size="Efficacy", color_continuous_scale="RdYlGn"))
            csv = df_manual.to_csv(index=False).encode("utf-8")
            excel = export_excel(df_manual)
            st.download_button("📥 Download Manual Results (CSV)", csv, "manual_results.csv", "text/csv")
            st.download_button("📘 Download Manual Results (Excel)", excel, "manual_results.xlsx")

    # ---------------- Final notes ----------------
    st.markdown("""
    **Notes & Caveats**:  
    - The ML surrogate is trained on a synthetic dataset that encodes domain heuristics — it is a fast **in-silico** proxy for screening, *not* a replacement for docking, ADMET profiling, or in-vitro/in-vivo validation.  
    - Preset compounds use approximate literature-informed values for illustrative comparison.  
    - Use exported CSV/Excel results for archiving, plotting or downstream modelling.
    """)






# --------------------------- VR HEADSET MODE ---------------------------
elif page == "VR Headset Mode 🕶️":
    import streamlit.components.v1 as components

    st.subheader("🧠 NeuroVR Headset Mode — Immersive Visualization")
    st.markdown("""
    This experimental mode enables **3D and VR visualization** of compound effects and stroke risk patterns.  
    Open the app via **HTTPS** on a **WebXR-compatible headset** (e.g., Oculus Quest, Pico, or HoloLens) to enter full VR.
    """)

    aframe_html = """
    <html>
      <head>
        <script src="https://aframe.io/releases/1.5.0/aframe.min.js"></script>
      </head>
      <body>
        <a-scene vr-mode-ui="enabled: true" embedded>
          <!-- Background -->
          <a-sky color="#ECECEC"></a-sky>

          <!-- Lighting -->
          <a-entity light="type: ambient; color: #BBB"></a-entity>
          <a-entity light="type: directional; color: #FFF; intensity: 0.8" position="1 1 0"></a-entity>

          <!-- Camera -->
          <a-entity position="0 1.6 0">
            <a-camera wasd-controls-enabled="true"></a-camera>
          </a-entity>

          <!-- Sample Drug Compounds (3D spheres) -->
          <a-entity id="compounds">
            <a-sphere position="-1 1.5 -3" radius="0.4" color="#EF2D5E" compound-name="Aspirin"></a-sphere>
            <a-sphere position="0 1.5 -4" radius="0.4" color="#4CC3D9" compound-name="Citicoline"></a-sphere>
            <a-sphere position="1 1.5 -3" radius="0.4" color="#7BC8A4" compound-name="tPA"></a-sphere>
            <a-sphere position="0 0.8 -2.5" radius="0.3" color="#FFC65D" compound-name="Edaravone"></a-sphere>
          </a-entity>

          <!-- Labels -->
          <a-text value="NeuroVR Lab — Compound Risk Space" color="#111" position="-1.2 2.5 -3"></a-text>
          <a-text value="(Look around or use headset controls to explore)" color="#333" position="-1.8 2.2 -3.2" width="4"></a-text>

          <!-- Animation Example -->
          <a-animation attribute="rotation" dur="12000" to="0 360 0" repeat="indefinite"></a-animation>

          <!-- Interaction Script -->
          <script>
            AFRAME.registerComponent('compound-info', {
              init: function () {
                this.el.addEventListener('click', () => {
                  const name = this.el.getAttribute('compound-name');
                  alert(`Compound: ${name}\\nBinding: variable\\nSolubility: variable\\nRisk: dynamic`);
                });
              }
            });

            // Attach info component to all compound spheres
            document.querySelectorAll('[compound-name]').forEach(e => e.setAttribute('compound-info', ''));
          </script>
        </a-scene>
      </body>
    </html>
    """

    components.html(aframe_html, height=600, scrolling=False)




# --------------------------- UPLOAD / RETRAIN DATASET ---------------------------
elif page == "Upload/Train Dataset 📂":
    st.subheader("Upload & Retrain Model (CSV must have: age,bp,cholesterol,glucose,bmi,stroke)")
    uploaded = st.file_uploader("Upload CSV", type=["csv"])
    if uploaded:
        user_df = pd.read_csv(uploaded)
        st.dataframe(user_df.head())
        if st.button("Retrain Model"):
            try:
                model, new_acc = train_model(user_df)
                st.success(f"Retrained model. New accuracy: {new_acc*100:.2f}%")
            except Exception as e:
                st.error(f"Retrain failed: {e}")

# --------------------------- RESEARCH REPORT ---------------------------
elif page == "Research Report 🧾":
    st.subheader("Generate PDF Report")
    name = st.text_input("Researcher Name", "Simon")
    institution = st.text_input("Institution", "NeuroVR Institute")
    if st.button("Generate PDF"):
        buffer = io.BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)
        c.setFont("Helvetica-Bold", 16)
        c.drawString(160, 800, "NeuroVR Stroke Research Report")
        c.setFont("Helvetica", 11)
        c.drawString(50, 760, f"Researcher: {name}")
        c.drawString(50, 740, f"Institution: {institution}")
        c.drawString(50, 720, f"Model accuracy (current session): {acc*100:.2f}%")
        c.drawString(50, 700, "Summary: AI-assisted stroke prediction and compound simulation.")
        c.showPage()
        c.save()
        st.download_button("Download Report PDF", buffer.getvalue(), file_name="NeuroVR_Report.pdf")

# --------------------------- FOOTER ---------------------------
st.markdown("---")
st.markdown(
    """
    <div style='text-align:center; color: gray; font-size: 13px;'>
        🧠 <b>System by Simon</b> | 📧 <a href='allinmer57@gmail.com'>allinmer57@gmail.com</a><br>
        © 2025 NeuroVR Lab — Experimental stroke research environment.
    </div>
    """,
    unsafe_allow_html=True
)
